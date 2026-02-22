"""
Load.py
Loads all 300 schemes, 20 categories, and 300 rules from BenefitBridge_Dataset.xlsx
Run AFTER schema.py has been executed.
Usage: python Load.py
"""
import mysql.connector
import openpyxl
import os

# â”€â”€ Dataset path â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
DATASET_PATH = r"D:\Academics\sem-3\Mini Project\Python+DBMS\coding\dataset\BenefitBridge_Dataset.xlsx"

if not os.path.exists(DATASET_PATH):
    print(f"âŒ File not found: {DATASET_PATH}")
    print("   Please check the path and filename.")
    exit(1)

print(f"ðŸ“‚ Using dataset: {DATASET_PATH}")

# â”€â”€ DB connection â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
db = mysql.connector.connect(
    host="localhost",
    user="root",
    passwd="2006",
    database="smart_beneficiary_system"
)
cursor = db.cursor()

try:
    wb = openpyxl.load_workbook(DATASET_PATH)
except Exception as e:
    print(f"âŒ Failed to open Excel file: {e}")
    exit(1)

cursor.execute("SET FOREIGN_KEY_CHECKS = 0")

# â”€â”€ 1. CATEGORIES (20 rows) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
print("\n[1/3] Loading Categories...")
cursor.execute("TRUNCATE TABLE Categories")
ws_cat = wb['Categories']
cat_count = 0
for row in ws_cat.iter_rows(min_row=2, values_only=True):
    cat_id, cat_name, desc = row[0], row[1], row[2]
    if cat_id is None:
        continue
    cursor.execute(
        "INSERT INTO Categories (category_id, category_name, description) VALUES (%s, %s, %s)",
        (cat_id, cat_name, desc)
    )
    cat_count += 1
db.commit()
print(f"  âœ…  {cat_count} categories loaded.")

# â”€â”€ 2. SCHEMES (300 rows) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Columns: scheme_id, scheme_name, target_category_id, category_name,
#          description, benefits, benefit_type, state, official_link, registration_link
print("\n[2/3] Loading Schemes...")
cursor.execute("TRUNCATE TABLE Schemes")
ws_sch = wb['Schemes']
sch_count = 0
for row in ws_sch.iter_rows(min_row=2, values_only=True):
    sid      = row[0]
    sname    = row[1]
    tcat     = row[2]
    # row[3] = category_name (skip)
    desc     = row[4]
    benefits = row[5]
    btype    = row[6]
    state    = row[7]
    official = row[8]
    reg_link = row[9] if len(row) > 9 else None

    if sid is None:
        continue

    cursor.execute(
        """INSERT INTO Schemes
           (scheme_id, scheme_name, target_category, description,
            benefits, benefit_type, state, official_link, registration_link)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (sid, sname, tcat, desc, benefits, btype, state,
         official if official else None,
         reg_link if reg_link else None)
    )
    sch_count += 1
db.commit()
print(f"  âœ…  {sch_count} schemes loaded.")

# â”€â”€ 3. RULE ENGINE (300 rows) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Columns: rule_id, scheme_id, scheme_name, category_id, category_name,
#          age_min, age_max, gender, location, min_income, max_income,
#          education_required, pension_status, disability_cert,
#          unemployment_status, business_turnover_limit
print("\n[3/3] Loading Rule Engine...")
cursor.execute("TRUNCATE TABLE Rule_Engine")
ws_rule = wb['Rule_Engine']
rule_count = 0
for row in ws_rule.iter_rows(min_row=2, values_only=True):
    rid       = row[0]
    sid       = row[1]
    # row[2]  = scheme_name (skip)
    cid       = row[3]
    # row[4]  = category_name (skip)
    a_min     = row[5]
    a_max     = row[6]
    gender    = row[7]
    loc       = row[8]
    min_inc   = row[9]
    max_inc   = row[10]
    edu       = row[11]
    pension   = row[12]
    disability= row[13]
    unemployed= row[14]
    turnover  = row[15]

    if rid is None:
        continue

    cursor.execute(
        """INSERT INTO Rule_Engine
           (rule_id, scheme_id, category_id,
            age_min, age_max, gender, location,
            min_income, max_income, education_required,
            pension_status, disability_cert,
            unemployment_status, business_turnover_limit)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (
            rid, sid, cid,
            a_min, a_max,
            gender     if gender     else None,
            loc        if loc        else None,
            min_inc, max_inc,
            edu        if edu        else None,
            1 if pension    else None,
            1 if disability else None,
            1 if unemployed else None,
            turnover
        )
    )
    rule_count += 1
db.commit()
print(f"  âœ…  {rule_count} rules loaded.")

cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
db.commit()

# â”€â”€ Final verification â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
print("\n" + "â”€" * 45)
print("ðŸ“Š  Final record counts in database:")
for table, label in [
    ("Categories", "Categories"),
    ("Schemes",    "Schemes   "),
    ("Rule_Engine","Rules     "),
]:
    cursor.execute(f"SELECT COUNT(*) FROM {table}")
    print(f"    {label} : {cursor.fetchone()[0]}")
print("â”€" * 45)

cursor.close()
db.close()
print("\nðŸŽ‰  Dataset loaded successfully! System is ready to use.")